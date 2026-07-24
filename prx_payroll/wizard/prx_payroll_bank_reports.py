import base64
import io
from openpyxl import Workbook
from odoo import api, fields, models
from ..models.configuration.prx_enum_selection import BankReports,BankTransactionReportType,SalaryType
from odoo.exceptions import UserError
from openpyxl.styles import numbers


class PRXPayrollBankReports(models.TransientModel):
    _name = 'prx.payroll.bank.reports'
    _description = 'Period bank reports wizard'

    period_id = fields.Many2one('prx.payroll.period', string="პერიოდი", required=True)
    bank = fields.Selection(BankReports.selection(), string='ბანკი', required=True)
    transaction_type = fields.Selection(BankTransactionReportType.selection(), string='ტრანზაქციის ტიპი', default='all',required=True)
    process_type = fields.Selection(SalaryType.selection(), string='პროცესის ტიპი',required=True)
    report_language = fields.Selection([
        ('ka_GE', 'Georgian'),
        ('en_US', 'English')
    ], string='ენა', default='ka_GE', required=True)
    file_download = fields.Binary("File", readonly=True)
    file_name = fields.Char("Filename")
    department_id = fields.Many2one("hr.department", string="დეპარტამენტი")

    def _get_transaction_domain(self):
        domain = [
            ('period_id', '=', self.period_id.id),
            ('code', '!=', False),
        ]
        if self.department_id:
            domain.append(('organization_unit_id', '=', self.department_id.id))
        if self.transaction_type == 'non_transferred':
            domain += [('transferred', '=', False)]
        elif self.transaction_type == 'transferred':
            domain += [('transferred', '=', True)]
        return domain

    def _get_employee_bank_account(self, employee):
        """თანამშრომლის სახელფასო ანგარიში Odoo-ს ვერსიისგან დამოუკიდებლად:
        primary_bank_account_id / bank_account_ids მხოლოდ ახალ ვერსიებში არსებობს"""
        if 'primary_bank_account_id' in employee._fields and employee.primary_bank_account_id:
            return employee.primary_bank_account_id
        if 'bank_account_ids' in employee._fields and employee.bank_account_ids:
            return employee.bank_account_ids[:1]
        return employee.bank_account_id or False

    def _get_employee_amounts_map(self):
        transaction_model = self.env['prx.payroll.transaction']
        domain = self._get_transaction_domain()
        transactions = transaction_model.search(domain)
        amounts_map = {}

        for transaction in transactions:
            employee_id = transaction.employee_id.id
            if not employee_id:
                continue

            employee_amounts = amounts_map.setdefault(employee_id, {
                'amount': 0.0,
                'salary_amount': 0.0,
                'advance_amount': 0.0,
                'bonus_amount': 0.0,
            })
            amount = transaction.amount or 0.0
            employee_amounts['amount'] += amount

            salary_type = transaction.worksheet_id.salary_type
            if salary_type == 'standard':
                employee_amounts['salary_amount'] += amount
            elif salary_type == 'avanse':
                employee_amounts['advance_amount'] += amount
            elif salary_type == 'one_time':
                employee_amounts['bonus_amount'] += amount

        return amounts_map

    def _get_employee_name(self, employee):
        if self.report_language == 'en_US':
            return f"{employee.prx_firstname_en or ''} {employee.prx_lastname_en or ''}".strip() or employee.name or ''
        return employee.name or ''

    def action_generate_bank_reports(self):
        self.ensure_one()
        wb = Workbook()
        ws = wb.active
        ws.title = "ბანკის ფაილი"
        if self.bank == 'bog':
            header, values = self.get_bog_excel_header_and_values()
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 45
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 18
            ws.column_dimensions['H'].width = 55
            ws.append(header)
        elif self.bank == 'tbc':
            header, values = self.get_tbc_excel_header_and_values()
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 20
            ws.append(header)
        else:
            raise UserError('აირჩიე ბანკი!')

        for value in values:
            ws.append(value)

        amount_columns = ['C', 'D', 'E', 'F'] if self.bank == 'tbc' else ['E', 'F', 'G']
        start_row = 3 if self.bank == 'tbc' else 2
        for row in range(start_row, ws.max_row + 1):
            for column in amount_columns:
                ws[f"{column}{row}"].number_format = numbers.FORMAT_NUMBER_00
            ws[f"A{row}"].number_format = numbers.FORMAT_TEXT

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        self.file_download = base64.b64encode(bio.read())
        self.file_name = f"Bank_Reports_{self.period_id.period+'_'+self.bank}.xlsx"
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }


    def get_bog_excel_header_and_values(self):
        if self.report_language == 'en_US':
            header = [
                'Receiver Account Number',
                'Receiver Bank Code (Optional)',
                'Receiver Name',
                'Purpose',
                'Transfer Amount',
                'Take Home Salary',
                'Take Home Advance',
                'Take Home Bonus',
                'Receiver Identification Code (Optional)'
            ]
            purpose = 'Salary'
        else:
            header = [
                'მიმღების ანგარიშის ნომერი',
                'მიმღები ბანკის კოდი(არასავალდებულო)',
                'მიმღების დასახელება',
                'დანიშნულება',
                'გადასარიცხი თანხა',
                'ხელზე ასაღები ხელფასი',
                'ხელზე ასაღები ავანსი',
                'ხელზე ასაღები ბონუსი',
                'მიმღების საიდენტიფიკაციო კოდი(არასავალდებულო)'
            ]
            purpose = 'ხელფასი'

        domain = self._get_transaction_domain() + [('worksheet_id.salary_type', '=', self.process_type)]

        txs = self.env['prx.payroll.transaction'].read_group(
            domain,
            ['employee_id','amount:sum(amount)',],
            ['employee_id'])

        values=[]
        employee_amounts_map = self._get_employee_amounts_map()
        employee_model = self.env['hr.employee']
        for empl in txs:
            employee = employee_model.browse(empl['employee_id'][0])
            bank_account = self._get_employee_bank_account(employee)
            employee_amounts = employee_amounts_map.get(employee.id, {})
            values.append([
                bank_account.acc_number if bank_account else '',
                bank_account.bank_id.bic if bank_account and bank_account.bank_id else '',
                self._get_employee_name(employee),
                purpose,
                empl['amount'] or 0.0,
                employee_amounts.get('salary_amount', 0.0),
                employee_amounts.get('advance_amount', 0.0),
                employee_amounts.get('bonus_amount', 0.0),
                employee.identification_id or '',
                ])
        return header, values

    def get_tbc_excel_header_and_values(self):
        if self.report_language == 'en_US':
            header = [
                'Account Number',
                "Employee's Name",
                'Amount',
                'Take Home Salary',
                'Take Home Advance',
                'Take Home Bonus',
                'Description',
            ]
            purpose = 'Salary'
        else:
            header = [
                'მიმღების ანგარიში',
                'მიმღების სახელი და გვარი',
                'თანხა',
                'ხელზე ასაღები ხელფასი',
                'ხელზე ასაღები ავანსი',
                'ხელზე ასაღები ბონუსი',
                'დანიშნულება',
            ]
            purpose = 'ხელფასი'

        domain = self._get_transaction_domain() + [('worksheet_id.salary_type', '=', self.process_type)]


        txs = self.env['prx.payroll.transaction'].read_group(
            domain,
            ['employee_id','amount:sum(amount)',],
            ['employee_id'])

        values=[]
        employee_amounts_map = self._get_employee_amounts_map()
        employee_model = self.env['hr.employee']
        for empl in txs:
            employee = employee_model.browse(empl['employee_id'][0])
            bank_account = self._get_employee_bank_account(employee)
            employee_amounts = employee_amounts_map.get(employee.id, {})
            values.append([
                bank_account.acc_number if bank_account else '',
                self._get_employee_name(employee),
                empl['amount'] or 0.0,
                employee_amounts.get('salary_amount', 0.0),
                employee_amounts.get('advance_amount', 0.0),
                employee_amounts.get('bonus_amount', 0.0),
                purpose,
                ])

        return header,values