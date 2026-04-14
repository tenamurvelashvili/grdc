import base64
import io
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
from odoo import fields, models
from odoo.exceptions import UserError


class PRXTransactionReport2(models.TransientModel):
    _name = 'prx.payroll.transaction.report.2'
    _description = 'Simplified Payroll Transaction Report'

    period_ids = fields.Many2many('prx.payroll.period', string="პერიოდი", required=True)
    salary_type_filter = fields.Selection(
        selection=[
            ('standard', 'ხელფასი (სტანდარტიული)'),
            ('avanse', 'ავანსი'),
            ('one_time', 'ერთჯერადი / ბონუსი'),
        ],
        string="ტრანზაქციის ტიპი",
    )
    file_download = fields.Binary("ფაილი", readonly=True)
    file_name = fields.Char("ფაილის სახელი")

    def generate_report(self):
        self.ensure_one()

        domain = [
            ('period_id', 'in', self.period_ids.ids),
            ('code', '!=', False),
            ('active', '=', True),
        ]
        if self.salary_type_filter:
            domain.append(('worksheet_id.salary_type', '=', self.salary_type_filter))

        txs = self.env['prx.payroll.transaction'].search(domain)
        if not txs:
            raise UserError('ჩანაწერი არ მოიძებნა!')

        # Pre-compute per-employee aggregates in a single pass (no extra DB queries)
        agg = defaultdict(lambda: {
            'ანაზღაურება': 0.0,
            'გადასახადი': 0.0,
            'დაქვითვა': 0.0,
            'ხელზე ასაღები თანხა': 0.0,
        })
        rows_by_key = {}

        for tr in txs:
            key = (tr.employee_id.id, tr.period_id.id)
            amount = tr.amount or 0.0
            agg[key]['ხელზე ასაღები თანხა'] += amount
            if tr.transaction_type == 'earning':
                agg[key]['ანაზღაურება'] += amount
            elif tr.transaction_type == 'tax':
                agg[key]['გადასახადი'] += amount
            elif tr.transaction_type == 'deduction':
                agg[key]['დაქვითვა'] += amount

            if key not in rows_by_key:
                rows_by_key[key] = {
                    'დეპარტამენტი': tr.employee_id.department_id.name or '',
                    'სახელი/გვარი': tr.employee_id.name or '',
                    'პირადი ნომერი': tr.employee_id.identification_id or '',
                    'პერიოდი': tr.period_id.period or '',
                    'ტაბელი': tr.worksheet_id.sequence or '',
                }

        # Merge aggregates and sort by department
        final_rows = []
        for key, base in rows_by_key.items():
            final_rows.append({**base, **agg[key]})
        final_rows.sort(key=lambda r: r.get('დეპარტამენტი') or '')

        # Build Excel workbook
        header = [
            'დეპარტამენტი', 'სახელი/გვარი', 'პირადი ნომერი', 'პერიოდი', 'ტაბელი',
            'ანაზღაურება', 'გადასახადი', 'დაქვითვა', 'ხელზე ასაღები თანხა',
        ]
        col_widths = [22, 25, 20, 15, 15, 15, 15, 15, 22]

        wb = Workbook()
        ws = wb.active
        ws.title = "ტრანზაქციები"

        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        header_fill = PatternFill(fill_type="solid", fgColor="FF8DB3E3")
        center = Alignment(horizontal="center", vertical="center")

        ws.append(header)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.alignment = center

        for row in final_rows:
            ws.append([
                row.get('დეპარტამენტი', ''),
                row.get('სახელი/გვარი', ''),
                row.get('პირადი ნომერი', ''),
                row.get('პერიოდი', ''),
                row.get('ტაბელი', ''),
                row.get('ანაზღაურება', 0.0),
                row.get('გადასახადი', 0.0),
                row.get('დაქვითვა', 0.0),
                row.get('ხელზე ასაღები თანხა', 0.0),
            ])

        # Numeric formatting starting from column 6 ('ანაზღაურება')
        numeric_start_col = 6
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(numeric_start_col, len(header) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    cell.value = float(cell.value)
                except (TypeError, ValueError):
                    cell.value = 0.0
                cell.number_format = numbers.FORMAT_NUMBER_00

        # Personal number column as text
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=3).number_format = numbers.FORMAT_TEXT

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        self.file_download = base64.b64encode(bio.read())
        self.file_name = "Transaction_Report_Simple.xlsx"

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }

