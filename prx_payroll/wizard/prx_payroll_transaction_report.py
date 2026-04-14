import base64
import io
from openpyxl import Workbook
from odoo import api, fields, models
from odoo.exceptions import UserError
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
from openpyxl.styles import numbers


class PRXPayrollTransactionReports(models.TransientModel):
    _name = 'prx.payroll.transaction.reports'
    _description = 'Period transaction reports wizard'

    period_ids = fields.Many2many('prx.payroll.period', string="პერიოდი", required=True)
    file_download = fields.Binary("File", readonly=True)
    file_name = fields.Char("Filename")

    def get_ordered_records(self):
        earning = self.env['prx.payroll.earning']
        tax = self.env['prx.payroll.tax']
        deduction = self.env['prx.payroll.deduction']

        earnings_nonzero = earning.search([('code', '!=', 0)], order='code asc')
        earnings_zero = earning.search([('code', '=', 0)])
        earnings_list = list(earnings_nonzero) + list(earnings_zero)

        taxes_nonzero = tax.search([('code', '!=', 0)], order='code asc')
        taxes_zero = tax.search([('code', '=', 0)])
        taxes_list = list(taxes_nonzero) + list(taxes_zero)

        deductions_nonzero = deduction.search([('code', '!=', 0)], order='code asc')
        deductions_zero = deduction.search([('code', '=', 0)])
        deductions_list = list(deductions_nonzero) + list(deductions_zero)

        return earnings_list + taxes_list + deductions_list

    def generate_transaction_report(self):
        self.ensure_one()
        dynamic_records = self.get_ordered_records()
        dynamic_header = []
        for rec in dynamic_records:
            if rec._name == 'prx.payroll.earning':
                dynamic_header.append(rec.earning)
            elif rec._name == 'prx.payroll.tax':
                dynamic_header.append(rec.tax)
            else:
                dynamic_header.append(rec.deduction)

        internal_keys = [f"{rec._name}_{rec.id}" for rec in dynamic_records]

        # fixed header columns (removed "ტაბელი") + dynamic headers
        header = [
            'დეპარტამენტი', 'სახელი/გვარი', 'პირადი ნომერი', 'პერიოდი','ტაბელი',
            'ანაზღაურება', 'გადასახადი', 'დაქვითვა',
            'ხელზე ასაღები თანხა', 'ხელზე ასაღები ხელფასი', 'ხელზე ასაღები ავანსი',
            'ხელზე ასაღები ბონუსი'
        ]
        # first row holds merged/dynamic headers for the dynamic part; create None placeholders
        header0 = [None] * len(header) + dynamic_header
        wb = Workbook()
        ws = wb.active
        ws.title = "ტრანზაქციები"
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15

        ws_color_header = PatternFill(fill_type="solid", fgColor="FF8DB3E3")
        center = Alignment(horizontal="center", vertical="center")

        ws.append(header0)
        ws.append(header)
        letters = [get_column_letter(i) for i in range(1, len(header0) + 1)]
        # style the first block of fixed headers (row 2)
        first_block_end = get_column_letter(len(header))  # column letter for the last fixed header (e.g. 'K')
        for cell in ws[f"A2:{first_block_end}2"][0]:
            cell.fill = ws_color_header
            cell.alignment = center
        # style the dynamic headers in row 1 (from the column after fixed headers to the last letter)
        if len(letters) > len(header):
            dynamic_start = get_column_letter(len(header) + 1)
            for cell in ws[f"{dynamic_start}1": f"{letters[-1]}1"][0]:
                cell.fill = ws_color_header
                cell.alignment = center
        else:
            # no dynamic columns
            pass

        for idx, col_letter in enumerate(letters, start=1):
            # widen dynamic columns and columns after the first fixed block
            if idx > len(header):
                ws.column_dimensions[col_letter].width = 25

        raw_rows = []
        txs = self.env['prx.payroll.transaction'].search([
            ('period_id', 'in', self.period_ids.ids),
            ('code', '!=', False),
            ('active','=',True),
        ])

        for tr in txs:
            # base row with required columns. Additional payroll-specific fields are left as placeholders
            # Now sums are filtered by both employee and period
            row = {
                'employee_id': tr.employee_id.id,
                'დეპარტამენტი': tr.employee_id.department_id.name or None,
                'სახელი/გვარი': tr.employee_id.name or None,
                'პირადი ნომერი': tr.employee_id.identification_id or None,
                'პერიოდი': tr.period_id.period or None,
                'ტაბელი': tr.worksheet_id.sequence or None,
                'ანაზღაურება': sum(tr.search([('employee_id','=',tr.employee_id.id),('code', '!=', False), ('period_id','=',tr.period_id.id), ('transaction_type','=','earning')]).mapped('amount')),
                'გადასახადი': sum(tr.search([('employee_id','=',tr.employee_id.id),('code', '!=', False), ('period_id','=',tr.period_id.id), ('transaction_type','=','tax')]).mapped('amount')),
                'დაქვითვა': sum(tr.search([('employee_id','=',tr.employee_id.id),('code', '!=', False), ('period_id','=',tr.period_id.id), ('transaction_type','=','deduction')]).mapped('amount')),
                # 'ჯამი':,
                # placeholders for the newly added columns (mock values)
                'ხელზე ასაღები თანხა': sum(tr.search([('employee_id','=',tr.employee_id.id),('code', '!=', False), ('period_id','=',tr.period_id.id)]).mapped('amount')),
                'ხელზე ასაღები ხელფასი': sum(tr.search([('employee_id','=',tr.employee_id.id),('code', '!=', False), ('period_id','=',tr.period_id.id), ('worksheet_id.salary_type','=','standard')]).mapped('amount')),
                'ხელზე ასაღები ავანსი': sum(tr.search([('employee_id','=',tr.employee_id.id),('code', '!=', False), ('period_id','=',tr.period_id.id), ('worksheet_id.salary_type','=','avanse')]).mapped('amount')),
                'ხელზე ასაღები ბონუსი': sum(tr.search([('employee_id','=',tr.employee_id.id),('code', '!=', False), ('period_id','=',tr.period_id.id), ('worksheet_id.salary_type','=','one_time')]).mapped('amount')),
            }
            for key in internal_keys:
                row[key] = pd.NA

            label_key = None
            if tr.transaction_type == 'earning' and tr.earning_id:
                for idx, rec in enumerate(dynamic_records):
                    if rec._name == 'prx.payroll.earning' and rec.id == tr.earning_id.id:
                        label_key = idx
                        break
            elif tr.transaction_type == 'tax' and tr.tax_id:
                for idx, rec in enumerate(dynamic_records):
                    if rec._name == 'prx.payroll.tax' and rec.id == tr.tax_id.id:
                        label_key = idx
                        break
            elif tr.transaction_type == 'deduction' and tr.deduction_id:
                for idx, rec in enumerate(dynamic_records):
                    if rec._name == 'prx.payroll.deduction' and rec.id == tr.deduction_id.id:
                        label_key = idx
                        break

            if label_key is not None:
                target_key = internal_keys[label_key]
                row[target_key] = tr.amount or 0.0

            raw_rows.append(row)

        df = pd.DataFrame(raw_rows)

        if df.empty:
            raise UserError('ჩანაწერი არ მოიძებნა!')

        for key in internal_keys:
            if key not in df.columns:
                df[key] = pd.NA

        for key in ['დეპარტამენტი', 'შიდა კოდი', 'პირადი ნომერი', 'პერიოდი','ტაბელი', 'ანაზღაურება', 'გადასახადი', 'დაქვითვა', 'ხელზე ასაღები თანხა', 'ხელზე ასაღები ხელფასი', 'ხელზე ასაღები ავანსი', 'ხელზე ასაღები ბონუსი']:
            if key not in df.columns:
                df[key] = pd.NA

        # aggregation for grouping. New fixed columns are included as first (placeholders)
        agg_dict = {
            'დეპარტამენტი': 'first',
            'სახელი/გვარი': 'first',
            'პირადი ნომერი': 'first',
            'ტაბელი': 'first',
            'ანაზღაურება': 'first',
            'გადასახადი': 'first',
            'დაქვითვა': 'first',
            # 'ჯამი': 'first',
            # placeholders for newly added columns
            'ხელზე ასაღები თანხა': 'first',
            'ხელზე ასაღები ხელფასი': 'first',
            'ხელზე ასაღები ავანსი': 'first',
            'ხელზე ასაღები ბონუსი': 'first',
        }
        for key in internal_keys:
            agg_dict[key] = 'sum'

        grouped_df = (
            df
            .groupby(['employee_id', 'პერიოდი'], as_index=False)
            .agg(agg_dict)
        )
        grouped_df = grouped_df.drop(columns=['employee_id'])

        present_internal = [k for k in internal_keys if k in grouped_df.columns]
        grouped_df[present_internal] = grouped_df[present_internal].replace(0.0, pd.NA)

        grouped_df = grouped_df.sort_values(by='დეპარტამენტი', ignore_index=True)

        for _, row in grouped_df.iterrows():
            excel_row = [
                row['დეპარტამენტი'],
                row['სახელი/გვარი'],
                row['პირადი ნომერი'],
                row['პერიოდი'],
                row['ტაბელი'],
                row['პერიოდი'],
                row['ანაზღაურება'],
                row['გადასახადი'] if 'გადასახადი' in row.index else row['გადასახადა'],
                row['დაქვითვა'],
                # placeholders (mock) for the 'hand' columns
                ('' if pd.isna(row['ხელზე ასაღები თანხა']) else row['ხელზე ასაღები თანხა']),
                ('' if pd.isna(row['ხელზე ასაღები ხელფასი']) else row['ხელზე ასაღები ხელფასი']),
                ('' if pd.isna(row['ხელზე ასაღები ავანსი']) else row['ხელზე ასაღები ავანსი']),
                ('' if pd.isna(row['ხელზე ასაღები ბონუსი']) else row['ხელზე ასაღები ბონუსი']),
            ] + [
                ('' if pd.isna(row[key]) else row[key])
                for key in internal_keys
            ]
            ws.append(excel_row)

        # numeric formatting: from the first numeric fixed column ('ანაზღაურება' -> col 6)
        start_col = len(['დეპარტამენტი', 'სახელი/გვარი', 'პირადი ნომერი', 'ტაბელი', 'პერიოდი']) + 1  # 6
        end_col = len(header) + len(internal_keys)

        for row_idx in range(3, ws.max_row + 1):
            for col_idx in range(start_col, end_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    cell.value = float(cell.value)
                except (TypeError, ValueError):
                    cell.value = 0.0
                cell.number_format = numbers.FORMAT_NUMBER_00

        for row in range(2, ws.max_row + 1):
            ws["{}{}".format("C", row)].number_format = numbers.FORMAT_TEXT
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        self.file_download = base64.b64encode(bio.read())
        self.file_name = "Transaction_Report.xlsx"

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }
