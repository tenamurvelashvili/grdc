import base64
import io
from openpyxl import Workbook
from odoo import api, fields, models
from odoo.exceptions import UserError
from openpyxl.styles import Font
from openpyxl.styles import numbers
import pandas as pd


class PRXPayrollPensionAlimonyReports(models.TransientModel):
    _name = 'prx.payroll.pension.alimony.reports'
    _description = 'Period tabel reports wizard'

    period_id = fields.Many2one('prx.payroll.period', string="პერიოდი", required=True)
    type = fields.Selection(selection=[('pension', 'საპენსიო'), ('alimony', 'ალიმენტი')], string="ტიპი")
    file_download = fields.Binary("File", readonly=True)
    file_name = fields.Char("Filename")

    def action_generate_report(self):
        self.ensure_one()
        bold = Font(size=12, bold=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "საპენსიო" if self.type == 'pension' else 'ალიმენტი'
        transaction = self.env['prx.payroll.transaction'].search([
            ('period_id', '=', self.period_id.id),
        ])
        header = ['კრედიტორი', 'კრედიტორის საიდენტიფიკაციო', 'კრედიტორის ანგარიში', 'თანხა']
        if self.type == 'pension':
            ws.append(header)
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 45
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 15

            for cell in ws["A1":"D1"][0]:
                cell.font = bold

            transaction = transaction.filtered(lambda tx: tx.deduction_id and tx.deduction_id.pension)

            if not transaction:
                raise UserError(" ტრანზაქციის ჩანაწერი არ მოიძებნა!")
            data = []
            for tx in transaction:
                cred = tx.creditor
                if not cred:
                    continue
                data.append({
                    'creditor_id': cred.id,
                    'creditor_name': cred.name or '',
                    'creditor_vat': cred.vat or '',
                    'acc_number': tx.creditor.bank_ids[0].acc_number if tx.creditor.bank_ids else "",
                    'amount': tx.amount if tx.amount * 2 else 0.0,
                })
            df = pd.DataFrame(data)
            if df.empty:
                raise UserError("პენსიის ჩანაწერი არ მოიძებნა!")
            grouped = (
                df
                .groupby(
                    ['creditor_id'],
                    as_index=False
                ).agg(
                    total_amount=('amount', 'sum'),
                    creditor_name=('creditor_name', 'first'),
                    creditor_vat=('creditor_vat', 'first'),
                    acc_number=('acc_number', 'first')
                )
            )

            for idx, row in grouped.iterrows():
                ws.append([
                    row['creditor_name'],
                    row['creditor_vat'],
                    row['acc_number'],
                    row['total_amount'],
                ])
            for row in range(2, ws.max_row + 1):
                ws["{}{}".format("D", row)].number_format = numbers.FORMAT_NUMBER_00

        if self.type == 'alimony':
            header.insert(0, 'თანამშრომელი')
            header.insert(1, 'თანამშრომელის საიდენტიფიკაციო')
            ws.append(header)

            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 45
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 40
            ws.column_dimensions['E'].width = 40
            ws.column_dimensions['F'].width = 15

            for cell in ws["A1":"F1"][0]:
                cell.font = bold
            transaction = transaction.filtered(
                lambda tx: tx.deduction_id and tx.deduction_id.alimony
            ).sorted(key=lambda x: x.employee_id.name or '')

            for rec in transaction:
                ws.append([
                    rec.employee_id.name or '',
                    rec.employee_id.identification_id or '',
                    rec.creditor.name or '',
                    rec.creditor.vat or '',
                    rec.creditor.bank_ids[0].acc_number if rec.creditor.bank_ids else '',
                    rec.amount if rec.amount * 2 else 0 or '',
                ])

            for row in range(2, ws.max_row + 1):
                ws["{}{}".format("F", row)].number_format = numbers.FORMAT_NUMBER_00
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        self.file_download = base64.b64encode(bio.read())
        self.file_name = f"{"Pension" if self.type == 'pension' else "Alimony"}_{self.period_id.period}_Report.xlsx"

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }
