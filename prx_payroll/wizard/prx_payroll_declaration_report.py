import base64
import io
from collections import defaultdict
import pandas as pd
from openpyxl import Workbook
from odoo import api, fields, models
from odoo.exceptions import UserError
from openpyxl.styles import numbers, Font, Alignment, PatternFill, Side
import logging
import pprint

_logger = logging.getLogger(__name__)


class PRXPayrollDeclarationWizard(models.TransientModel):
    _name = 'prx.payroll.declaration.wizard'
    _description = 'Payroll declaration report wizard'

    period_id = fields.Many2one('prx.payroll.period', string="პერიოდი", required=True)
    file_download = fields.Binary("File", readonly=True)
    file_name = fields.Char("Filename")

    def action_generate_declaration(self):
        self.ensure_one()

        txs = self.env['prx.payroll.transaction'].search([
            ('period_id', '=', self.period_id.id),
        ])

        employee_txs_cache = {}
        data = []
        buckets_by_employee = defaultdict(list)
        processed_keys = set()

        for t in txs:
            if t.transaction_type != 'earning' or not t.earning_id or not t.employee_id:
                continue

            earning_tax_report = t.earning_id.tax_report
            if not earning_tax_report:
                continue

            emp_id = t.employee_id.id
            transferred_time = t.prx_transferred_time
            transferred_time_key = str(transferred_time) if transferred_time else '__no_transfer__'
            key = (emp_id, earning_tax_report.id, transferred_time_key)
            if key in processed_keys:
                continue
            processed_keys.add(key)

            if emp_id not in employee_txs_cache:
                employee_txs_cache[emp_id] = txs.filtered(lambda r, employee=emp_id: r.employee_id.id == employee)
            employee_txs = employee_txs_cache[emp_id]

            earning_amount = sum(
                r.amount
                for r in employee_txs
                if (
                        r.transaction_type == 'earning'
                        and r.earning_id
                        and r.earning_id.tax_report == earning_tax_report
                        and r.prx_transferred_time == transferred_time
                )
            )

            tax_base_amount = sum(
                base.amount
                for base in employee_txs
                if base.include_tax_base
                and base.prx_transferred_time == transferred_time
                and (
                        (base.earning_id and base.earning_id.tax_report == earning_tax_report)
                        or (not base.earning_id and not earning_tax_report)
                )
            )

            pension_sum_proportion = sum(
                r.pension_proportion
                for r in employee_txs
                if (
                        r.transaction_type == 'earning'
                        and r.prx_transferred_time == transferred_time
                        and (
                                (r.earning_id and r.earning_id.tax_report == earning_tax_report)
                                or (not r.earning_id and not earning_tax_report)
                        )
                )
            )

            tax_sum_proportion = sum(
                r.tax_proportion
                for r in employee_txs
                if (
                        r.transaction_type == 'earning'
                        and r.prx_transferred_time == transferred_time
                        and (
                                (r.earning_id and r.earning_id.tax_report == earning_tax_report)
                                or (not r.earning_id and not earning_tax_report)
                        )
                )
            )
            _logger.info(
                f"EMPLOYEE {emp_id} EARNING_AMOUNT: {earning_amount} | TAX_BASE_AMOUNT: {tax_base_amount} | PENSION_PROP: {pension_sum_proportion} | TAX_PROP: {tax_sum_proportion}")
            tax_code = earning_tax_report.code or ''
            info = self.env['prx.payroll.report']._get_employee_create_vals(t.employee_id)
            gross_rate = t.tax_id.rate_gross * 100 if t.tax_id else 0.0
            pension_proportion = t.pension_proportion or 0.0
            tax_proportion = t.tax_proportion or 0.0

            entry = {
                'employee_id': emp_id,
                'period_id': self.period_id.id,
                'personal_number': t.personal_number or '',
                'first_name': t.employee_id.first_name,
                'last_name': t.employee_id.last_name,
                'private_street': info['private_street'],
                'resident_country': t.employee_id.tax_country.code or '',
                'tax_category': t.employee_id.tax_category.code or '',
                'tax_report': tax_code,
                'amount': t.amount or 0.0,
                # here bullshit fuck
                'rate_gross': gross_rate,
                'earning_amount': earning_amount,
                'another_benefit': (earning_amount - abs(pension_sum_proportion)),
                'tax_proportion': tax_sum_proportion,
                '_tax_base_amount': tax_base_amount,
                'payment_date': t.prx_transferred_time if t.prx_transferred_time else t.period_id.payment_date,
                'transferred_time': transferred_time_key,
            }
            data.append(entry)
            _logger.info(f"ENTRY: {pprint.pformat(entry)}")
            buckets_by_employee[emp_id].append(entry)

        # _logger.info(f"BUCKETS: {buckets_by_employee}")
        # Tax base tax calculation logic:

        _logger.info(f"BUCKETTSTSTSTTTSS BY EMPLOYEEE: {pprint.pformat(buckets_by_employee)}")

        # update

        for employee_id, buckets in buckets_by_employee.items():
            employee_taxes = employee_txs_cache.get(employee_id, self.env['prx.payroll.transaction'])

            # get ones that has tax base
            tax_records = employee_taxes.filtered(
                lambda r: r.transaction_type == 'tax'
                          and r.tax_id
                          and r.tax_id.rate_base > 0
            )
            if not tax_records:
                continue

            sorted_buckets = sorted(buckets, key=lambda b: b['_tax_base_amount'], reverse=True)
            sorted_taxes = sorted(tax_records, key=lambda rec: abs(rec.amount or 0.0), reverse=True)

            counter = 0
            for bucket, tax in zip(sorted_buckets, sorted_taxes):
                # # _logger.info(f"PROCCESSING BUCKET: {bucket} WITH TAX: {tax}")
                # counter += 1
                # _logger.info(f"LOOP COUNTER: {counter}")
                # base_amount = bucket['_tax_base_amount'] or 0.0
                # benefit = 0.0
                # if tax.amount == 0:
                #     benefit = base_amount
                # elif tax.tax_id.rate_gross:
                #     benefit = (tax.amount / tax.tax_id.rate_gross) + base_amount

                # bucket['another_benefit'] = benefit
                bucket['rate_gross'] = tax.tax_id.rate_gross * 100 if tax.tax_id and bucket['rate_gross'] == 0 else 0.0
                # _logger.info(
                #     "MAPPED TAX %s (amount=%s) TO DECLARATION BUCKET tax_report=%s BASE=%s -> BENEFIT=%s",
                #     tax.id,
                #     tax.amount,
                #     bucket['tax_report'],
                #     base_amount,
                #     benefit,
                # )

        # Format the FINAL DATA log for better readability

        self._update_bucket_taxes(buckets_by_employee, self.period_id)
        _logger.info("FINAL DATA:\n%s", pprint.pformat(data))

        # update: tax gross_rates, recheck for each

        df = pd.DataFrame(data)
        if df.empty:
            raise UserError('ჩანაწერი არ მოიძებნა!')
        if '_tax_base_amount' in df.columns:
            df = df.drop(columns=['_tax_base_amount'])

        grouped = (
            df.groupby(['employee_id', 'period_id', 'tax_report', 'transferred_time'], dropna=False, as_index=False)
            .agg(
                earning_amount=('earning_amount', 'first'),
                first_name=('first_name', 'first'),
                last_name=('last_name', 'first'),
                tax_report=('tax_report', 'first'),
                private_street=('private_street', 'first'),
                resident_country=('resident_country', 'first'),
                tax_category=('tax_category', 'first'),
                rate_gross=('rate_gross', 'first'),
                personal_number=('personal_number', 'first'),
                another_benefit=('another_benefit', 'first'),
                payment_date=('payment_date', 'first')
            )
            .reset_index()
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "საშემოსავლო დეკლარაცია"

        header = [
            "საიდენტიფიკაციო ნომერი (პირადი ნომერი)",
            "თანხის მიმღების სახელი/სამართლებრივი ფორმა",
            "თანხის მიმღების გვარი/დასახელება",
            "მისამართი",
            "პირის რეზიდენტობა (ქვეყანა)",
            "შემოსავლის მიმღებ პირთა კატეგორია",
            "განაცემის სახე",
            "განაცემი თანხა (ლარი)",
            "სხვა შეღავათი",
            "გაცემის თარიღი",
            "წყაროსთან დასაკავებელი გადასახადის განაკვეთი",
            "საერთაშორისო ხელშეკრულების საფუძველზე გათავისუფლებას დაქვემდებარებული გადასახადის თანხა (ლარი)",
            "ორმაგი დაბეგვრის თავიდან აცილების შესახებ ხელშეკრულების საფუძველზე ჩათვლას დაქვემდებარებული, უცხო ქვეყანაში გადახდილი გადასახადის თანხა / შესამცირებელი საშემოსავლო გადასახადი (ლარი)"
        ]
        ws.append(header)
        header_fill = PatternFill(fill_type="solid", fgColor="EEECE1")
        thin_side = Side(border_style="thin", color="000000")
        for col in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(size=11)
            cell.fill = header_fill
            cell.border = thin_side
            cell.alignment = Alignment(
                wrap_text=True,
                horizontal="center",
                vertical="center",
                textRotation=0
            )
            # 3) ვრცელი სვეტის სიგანე
            letter = cell.column_letter
            ws.column_dimensions[letter].width = 25

        ws.row_dimensions[1].height = 168

        ws.freeze_panes = "A2"

        for _, r in grouped.iterrows():
            ws.append([
                r['personal_number'],
                r['first_name'],
                r['last_name'],
                r['private_street'],
                r['resident_country'],
                r['tax_category'],
                r['tax_report'],
                r['earning_amount'],
                r['another_benefit'],
                r['payment_date'],
                r['rate_gross'],
                0,
                0,
            ])

        for row in range(2, ws.max_row + 1):
            ws["{}{}".format("H", row)].number_format = numbers.FORMAT_NUMBER_00
            ws["{}{}".format("I", row)].number_format = numbers.FORMAT_NUMBER_00
            ws["{}{}".format("M", row)].number_format = numbers.FORMAT_NUMBER_00
            # ws["{}{}".format("J", row)].number_format = numbers.FORMAT_DATE_DDMMYY

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        self.file_download = base64.b64encode(bio.read())
        self.file_name = f"Declaration_{self.period_id.period}.xlsx"

        _logger.info("\n--- Declaration Report Generated ---\n")
        for _, row in grouped.iterrows():
            _logger.info(
                "Declaration row | employee=%s | tax_report=%s | earning=%s | benefit=%s",
                row['employee_id'],
                row['tax_report'],
                row['earning_amount'],
                row['another_benefit'],
            )
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }

    def _update_bucket_taxes(self, buckets_by_employee, period):
        for emp_id in buckets_by_employee.keys():
            # this should only be one
            employee_tx = self.env['prx.payroll.transaction'].search([
                ('employee_id', '=', emp_id),
                ('period_id', '=', period.id),
                ('transaction_type', '=', 'tax'),
                # ('tax_id.rate_base', '>', 0)
            ], limit=1)

            _logger.info(
                f"EMPLOYEE {emp_id} TAX TRANSACTION: {employee_tx} WITH TAX_ID: {employee_tx.tax_id if employee_tx else 'NO_TX'}")
            for bucket in buckets_by_employee[emp_id]:
                rate_gross = employee_tx.tax_id.rate_gross if employee_tx and employee_tx.tax_id else 0.0
                _logger.info(f"RATE GROSS FOR EMPLOYEE FOUND {emp_id}: {rate_gross}")
                bucket['rate_gross'] = rate_gross * 100 if bucket['rate_gross'] == 0 else bucket['rate_gross']
                bucket['another_benefit'] = (bucket['another_benefit']) - abs(bucket['tax_proportion'] / rate_gross)

                # lord help us
                if bucket['another_benefit'] <= 0.1:
                    bucket['another_benefit'] = 0.0

