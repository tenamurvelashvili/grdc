import base64
import io
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter
from odoo import fields, models
from odoo.exceptions import UserError


class PRXTransactionPivotReport(models.TransientModel):
    _name = 'prx.payroll.transaction.pivot.report'
    _description = 'Payroll Transaction Pivot Style Report'

    period_ids = fields.Many2many('prx.payroll.period', string="პერიოდი", required=True)
    salary_type_filter = fields.Selection(
        selection=[
            ('standard', 'ხელფასი (სტანდარტიული)'),
            ('avanse', 'ავანსი'),
            ('one_time', 'ერთჯერადი / ბონუსი'),
        ],
        string="ტრანზაქციის ტიპი",
    )
    file_download = fields.Binary("ფაილი", readonly=True)
    file_name = fields.Char("ფაილის სახელი")

    def _column_info(self, tr):
        """Return (code, label) of the dynamic column a transaction belongs to."""
        if tr.transaction_type == 'earning':
            rec = tr.earning_id
            name = rec.earning if rec else ''
        elif tr.transaction_type == 'tax':
            rec = tr.tax_id
            name = rec.tax if rec else ''
        else:
            rec = tr.deduction_id
            name = rec.deduction if rec else ''
        code = (rec.code if rec and rec.code else tr.code) or ''
        return code, (name or code)

    def generate_report(self):
        self.ensure_one()

        domain = [
            ('period_id', 'in', self.period_ids.ids),
            ('transaction_type', 'in', ['earning', 'tax', 'deduction']),
            ('code', '!=', False),
            ('active', '=', True),
        ]
        if self.salary_type_filter:
            domain.append(('worksheet_id.salary_type', '=', self.salary_type_filter))

        txs = self.env['prx.payroll.transaction'].search(domain)
        if not txs:
            raise UserError('ჩანაწერი არ მოიძებნა!')

        # Collect the dynamic columns actually present in the filtered transactions,
        # grouped by transaction type, ordered by code
        group_order = ['earning', 'tax', 'deduction']
        group_labels = {
            'earning': 'დარიცხვები',
            'tax': 'გადასახადები',
            'deduction': 'დაქვითვები',
        }
        found_columns = {t: set() for t in group_order}
        rows_by_key = {}
        amounts = defaultdict(lambda: defaultdict(float))

        for tr in txs:
            col = self._column_info(tr)
            found_columns[tr.transaction_type].add(col)

            key = (tr.employee_id.id, tr.period_id.id)
            if key not in rows_by_key:
                rows_by_key[key] = {
                    'დეპარტამენტი': tr.employee_id.department_id.name or '',
                    'სახელი/გვარი': tr.employee_id.name or '',
                    'პირადი ნომერი': tr.employee_id.identification_id or '',
                    'პერიოდი': tr.period_id.period or '',
                    'ტაბელი': tr.worksheet_id.sequence or '',
                }
            amounts[key][(tr.transaction_type,) + col] += tr.amount or 0.0

        # (code, label) tuples ordered by code inside each group; empty codes last
        ordered_columns = {
            t: sorted(found_columns[t], key=lambda c: (c[0] == '', c[0], c[1]))
            for t in group_order
        }
        dynamic_keys = [
            (t,) + col
            for t in group_order
            for col in ordered_columns[t]
        ]

        fixed_header = ['დეპარტამენტი', 'სახელი/გვარი', 'პირადი ნომერი', 'პერიოდი', 'ტაბელი']
        n_fixed = len(fixed_header)
        total_col_idx = n_fixed + len(dynamic_keys) + 1

        wb = Workbook()
        ws = wb.active
        ws.title = "ტრანზაქციები"

        header_fill = PatternFill(fill_type="solid", fgColor="FF8DB3E3")
        group_fills = {
            'earning': PatternFill(fill_type="solid", fgColor="FFC6E0B4"),
            'tax': PatternFill(fill_type="solid", fgColor="FFF8CBAD"),
            'deduction': PatternFill(fill_type="solid", fgColor="FFFFE699"),
        }
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        bold_font = Font(bold=True)
        thin = Side(style='thin')
        box = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Header row 1: fixed columns merged over rows 1-2, one merged cell per
        # transaction type group, 'ჯამი' merged over rows 1-2
        for col_idx, title in enumerate(fixed_header, start=1):
            ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.fill = header_fill
            cell.alignment = center
            cell.font = bold_font

        col_cursor = n_fixed + 1
        for t in group_order:
            group_cols = ordered_columns[t]
            if not group_cols:
                continue
            start = col_cursor
            end = col_cursor + len(group_cols) - 1
            if end > start:
                ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
            cell = ws.cell(row=1, column=start, value=group_labels[t])
            cell.fill = group_fills[t]
            cell.alignment = center
            cell.font = bold_font
            # Header row 2: individual code columns
            for offset, (code, label) in enumerate(group_cols):
                sub_cell = ws.cell(row=2, column=start + offset, value=label)
                sub_cell.fill = group_fills[t]
                sub_cell.alignment = center
                ws.column_dimensions[get_column_letter(start + offset)].width = 18
            col_cursor = end + 1

        ws.merge_cells(start_row=1, start_column=total_col_idx, end_row=2, end_column=total_col_idx)
        total_header = ws.cell(row=1, column=total_col_idx, value='ჯამი')
        total_header.fill = header_fill
        total_header.alignment = center
        total_header.font = bold_font

        for row_idx in (1, 2):
            for col_idx in range(1, total_col_idx + 1):
                ws.cell(row=row_idx, column=col_idx).border = box

        for i, w in enumerate([22, 25, 20, 15, 15], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.column_dimensions[get_column_letter(total_col_idx)].width = 18

        # Data rows sorted by department
        final_rows = sorted(
            rows_by_key.items(),
            key=lambda item: (item[1]['დეპარტამენტი'], item[1]['სახელი/გვარი']),
        )
        for key, base in final_rows:
            row_amounts = amounts[key]
            dynamic_values = [row_amounts.get(k, 0.0) for k in dynamic_keys]
            ws.append([
                base['დეპარტამენტი'],
                base['სახელი/გვარი'],
                base['პირადი ნომერი'],
                base['პერიოდი'],
                base['ტაბელი'],
            ] + dynamic_values + [sum(dynamic_values)])

        # Numeric formatting on dynamic and total columns
        numeric_start_col = n_fixed + 1
        for row_idx in range(3, ws.max_row + 1):
            for col_idx in range(numeric_start_col, total_col_idx + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    cell.value = float(cell.value)
                except (TypeError, ValueError):
                    cell.value = 0.0
                cell.number_format = numbers.FORMAT_NUMBER_00

        # Personal number column as text
        for row_idx in range(3, ws.max_row + 1):
            ws.cell(row=row_idx, column=3).number_format = numbers.FORMAT_TEXT

        # Totals row
        last_data_row = ws.max_row
        thick = Side(style='medium')
        for col_idx in range(1, total_col_idx + 1):
            cell = ws.cell(row=last_data_row, column=col_idx)
            existing = cell.border
            cell.border = Border(
                left=existing.left,
                right=existing.right,
                top=existing.top,
                bottom=thick,
            )

        totals_row = ['ჯამი', '', '', '', '']
        for col_idx in range(numeric_start_col, total_col_idx + 1):
            totals_row.append(sum(
                ws.cell(row=r, column=col_idx).value or 0
                for r in range(3, last_data_row + 1)
            ))
        ws.append(totals_row)

        totals_row_idx = ws.max_row
        for col_idx in range(1, total_col_idx + 1):
            cell = ws.cell(row=totals_row_idx, column=col_idx)
            cell.font = bold_font
            if col_idx >= numeric_start_col:
                cell.number_format = numbers.FORMAT_NUMBER_00

        ws.freeze_panes = "A3"

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        self.file_download = base64.b64encode(bio.read())
        self.file_name = "Transaction_Report_Pivot.xlsx"

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }
