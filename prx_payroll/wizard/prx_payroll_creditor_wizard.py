import base64
import io
from openpyxl import Workbook
from odoo import api, fields, models
from odoo.exceptions import UserError
from openpyxl.styles import Font,Alignment,numbers

class PrxPayrollCreditorReports(models.TransientModel):
    _name = "prx.payroll.creditor.reports"
    _description = "Prx Payroll Creditor Wizard"

    period_id = fields.Many2one('prx.payroll.period',string='პერიოდი')
    transactions_creditors = fields.Many2many('res.partner',compute="_compute_transaction_creditors")
    creditor = fields.Many2one('res.partner',string='კრედიტორი')
    file_download = fields.Binary("File", readonly=True)
    file_name = fields.Char("Filename")

    @api.depends('period_id')
    def _compute_transaction_creditors(self):
        for rec in self:
            if rec.period_id:
                txs = self.env['prx.payroll.transaction'].search([
                    ('period_id', '=', rec.period_id.id)
                ])
                creditors = txs.mapped('creditor')
                rec.transactions_creditors = [(6, 0, creditors.ids)]
            else:
                rec.transactions_creditors = [(5,)]


    def generate_creditor_report(self):
        self.ensure_one()
        bold = Font(size=12, bold=True)
        wb = Workbook()
        ws = wb.active
        ws.title = 'კრედიტორის რეპორტი'
        transactions = self.env['prx.payroll.transaction'].search([
            ('period_id', '=', self.period_id.id),
            ('creditor', '!=', False),
        ]).sorted(key=lambda x: x.creditor.name or '')
        if self.creditor:
            transactions = transactions.filtered(lambda tx: tx.creditor.id == self.creditor.id)

        header = ['თანამშრომელი','თანამშრომელის საიდენტიფიკაციო','კრედიტორი', 'კრედიტორის საიდენტიფიკაციო', 'კრედიტორის ანგარიში', 'თანხა']
        ws.append(header)

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 15

        for cell in ws["A1":"F1"][0]:
            cell.font = bold

        for rec in transactions:
            ws.append([
                rec.employee_id.name or '',
                rec.employee_id.identification_id or '',
                rec.creditor.name or '',
                rec.creditor.vat or '',
                rec.creditor.bank_ids[0].acc_number if rec.creditor.bank_ids else '',
                rec.amount if rec.amount else '',
            ])
        total_amount = round(sum(rec.amount for rec in transactions),2)
        ws.append(['', '', '', '', '', ''])
        ws.append(['', '', '', '', '', ''])
        ws.append(['', '', '', '', 'ჯამური თანხა:', total_amount])
        total_row = ws.max_row
        ws[f'E{total_row}'].font = bold
        ws[f'F{total_row}'].font = bold
        ws.cell(row=total_row, column=1).font = bold
        ws.cell(row=total_row, column=6).font = bold
        for row in range(2, ws.max_row + 1):
            ws["{}{}".format("F", row)].number_format = numbers.FORMAT_NUMBER_00

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        self.file_download = base64.b64encode(bio.read())
        self.file_name = f"Creditor_{self.period_id.period}_Report.xlsx"

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }