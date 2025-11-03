import base64
import io
from openpyxl import Workbook
from odoo import api, fields, models
from odoo.exceptions import UserError
from openpyxl.styles import Font
from openpyxl.styles import numbers

georgian_headers = [
    "ორგანიზაციული ერთეული",
    "პოზიცია",
    "დოკუმენტის ნომერი",
    "სახელი და გვარი",
    "საიდენთიფიკაციო კოდი",
    "პერიოდი",
    "სტატუსი",
    "ანაზღაურება",
    "თანხა",
]

english_headers = [
    "Organization Unit",
    "Position",
    "Document Number",
    "Employee's Name",
    "Identification Number",
    "Period",
    "Status",
    "Earning Name",
    "Amount",
]


class PRXPayrollTabelReports(models.TransientModel):
    _name = 'prx.payroll.tabel.reports'
    _description = 'Period tabel reports wizard'

    period_ids = fields.Many2many('prx.payroll.period', string="პერიოდი", required=True)
    department_ids = fields.Many2many('hr.department', string="დეპარტამენტი")

    file_download = fields.Binary("File", readonly=True)
    file_name = fields.Char("Filename")

    def action_generate_tabel_report(self):
        self.ensure_one()
        blue_font = Font(color="FF4848E8", size=12, bold=True)
        bold = Font(size=12, bold=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "ტრანზაქციები"
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 15
        ws.append(georgian_headers)
        ws.append(english_headers)

        for cell in ws["A1":"I1"][0]:
            cell.font = blue_font
        for cell in ws["A2":"I2"][0]:
            cell.font = bold

        tabel = self.env['prx.payroll.worksheet'].search([
            ('period_id', 'in', self.period_ids.ids)
        ]).sorted(key=lambda x: x.worker_id.department_id.name or '')
        if self.department_ids:
            employee_ids = [rec.worker_id.id for rec in tabel if rec.worker_id.department_id in self.department_ids]
            tabel = tabel.filtered(lambda x: x.worker_id.id in employee_ids)

        last_dept = None
        for rec in tabel:
            dept = rec.worker_id.department_id.name or ''
            if dept != last_dept:
                dept_to_write = dept
                last_dept = dept
            else:
                dept_to_write = ''
            totals_amount = {}
            for earn in rec.worksheet_detail_ids:
                key = earn.earning_id.earning_id.earning
                if key not in totals_amount:
                    totals_amount[key] = 0.0
                totals_amount[key] += earn.amount
            earning_dict = [
                {"earning": desc, "amount": total}
                for desc, total in totals_amount.items()
            ]
            for earn in earning_dict:
                ws.append([
                    dept_to_write,
                    rec.worker_id.job_id.name or '',
                    rec.sequence or '',
                    rec.worker_id.name or '',
                    rec.worker_id.identification_id or '',
                    rec.period_id.period or '',
                    rec.status or '',
                    earn.get('earning') or '',
                    earn.get('amount') or 0.0,
                ])
        for row in range(2, ws.max_row + 1):
            ws["{}{}".format("E", row)].number_format = numbers.FORMAT_TEXT
            ws["{}{}".format("I", row)].number_format = numbers.FORMAT_NUMBER_00

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        self.file_download = base64.b64encode(bio.read())
        self.file_name = f"Tabel_Report.xlsx"

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }
