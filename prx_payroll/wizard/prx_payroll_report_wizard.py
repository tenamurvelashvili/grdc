from odoo import models, fields
import io
import json
import base64
from openpyxl.styles import Font
from openpyxl.styles import Alignment, PatternFill
from openpyxl import Workbook


class PRXPayrollReport(models.TransientModel):
    _name = 'prx.payroll.report'
    _description = 'Payroll report wizard'

    company_id = fields.Many2one('res.company', string='კომპანია', default=lambda self: self.env.company, required=True)
    type = fields.Selection([('xlsx', 'XLSX'), ('pdf', 'PDF')], required=True, default='pdf')
    language = fields.Selection([('us', 'English'), ('geo', 'Georgian')], required=True, default='geo')
    worksheet = fields.Many2one('prx.payroll.worksheet', string="Worksheet")
    reports = fields.Selection([('payslip_xslx', 'საშემოსავლოს უწყისი'),
                                ('tax_declaration', 'საშემოსავლოს დეკლარაცია')], default='payslip_xslx')
    file_download = fields.Binary("File", readonly=True)
    file_name = fields.Char("Filename")

    def _get_employee_create_vals(self, employee_id):
        self.ensure_one()
        address_id = employee_id.address_id.id
        address_sudo = self.env['res.partner'].sudo().browse(address_id)
        return {
            'name': employee_id.name,
            'private_street': (
                                  f"{employee_id.private_country_id.name}, " if employee_id.private_country_id else ""
                              ) + f"{employee_id.private_city or ''}, {employee_id.private_street or employee_id.private_street2 or ''}",
            'private_city': address_sudo.city,
            'private_state_id': address_sudo.state_id.id,
            'private_zip': address_sudo.zip,
            'private_country_id': address_sudo.country_id.id,
            'private_phone': address_sudo.phone,
            'work_email': employee_id.work_email,
            'address_id': self.env.company.partner_id.street or '',
            'phone': employee_id.work_phone or '',
        }

    def generate_tax_declaration(self):
        datas = [
            {"id": 1, "name": "საიდენტიფიკაციო ნომერი (პირადი ნომერი)", "value": 0},
            {"id": 2, "name": "თანხის მიმღების სახელი/სამართლებრივი ფორმა", "value": 0},
            {"id": 3, "name": "თანხის მიმღების გვარი/დასახელება", "value": 0},
            {"id": 4, "name": "მისამართი"},
            {"id": 5, "name": "პირის რეზიდენტობა (ქვეყანა)", "value": 0},
            {"id": 6, "name": "შემოსავლის მიმღებ პირთა კატეგორია", "value": 0},
            {"id": 7, "name": "განაცემის სახე", "value": 0},
            {"id": 8, "name": "განაცემი თანხა (ლარი)", "value": 0},
            {"id": 9,
             "name": "დაგროვებითი პენსიის შესახებ კანონის შესაბამისად დასაქმებულის სახელითა და ხარჯით განხორციელებული საპენსიო შენატანი",
             "value": 0},
            {"id": 10, "name": "სხვა შეღავათი", "value": 0},
            {"id": 11, "name": "გაცემის თარიღი", "value": 0},
            {"id": 12, "name": "წყაროსთან დასაკავებელი გადასახადის განაკვეთი", "value": 0},
            {"id": 13, "name": "წყაროსთან დაკავებული გადასახადი (ლარი)", "value": 0},
            {"id": 14,
             "name": "საერთაშორისო ხელშეკრულების საფუძველზე გათავისუფლებას დაქვემდებარებული გადასახადის თანხა (ლარი)",
             "value": 0},
            {"id": 15,
             "name": "ორმაგი დაბეგვრის თავიდან აცილების შესახებ ხელშეკრულების საფუძველზე ჩათვლას დაქვემდებარებული, უცხო ქვეყანაში გადახდილი გადასახადის თანხა / შესამცირებელი საშემოსავლო გადასახადი (ლარი)",
             "value": 0}
        ]
        wb = Workbook()
        ws = wb.active
        ws.title = "Tax Declaration"

        header = [item['name'] for item in datas]
        ws.append(header)

        values = [item.get('value', '') for item in datas]
        ws.append(values)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        data = buf.read()
        buf.close()

        self.file_download = base64.b64encode(data)
        self.file_name = 'tax_declaration.xlsx'

    def generate(self):
        self.ensure_one()

        w = self.worksheet
        transactions = self.env['prx.payroll.transaction'].search([('worksheet_id', '=', w.id)])
        employee = self._get_employee_create_vals(w.worker_id)

        earning_tx = transactions.filtered(lambda t: t.transaction_type == 'earning')
        tax_tx = transactions.filtered(lambda t: t.transaction_type == 'tax')
        deduction_tx = transactions.filtered(lambda t: t.transaction_type == 'deduction')

        earning_ids = [{'code': t.code, 'qty': t.qty, 'rate': t.rate, 'amount': t.amount} for t in earning_tx]
        tax_ids = [{'code': t.code, 'rate': t.rate, 'amount': t.amount} for t in tax_tx]
        deduction_ids = [{'code': t.code, 'rate': t.rate, 'amount': t.amount} for t in deduction_tx]

        earning_total = sum(earning_tx.mapped('amount'))
        tax_total = sum(tax_tx.mapped('amount'))
        deduction_total = sum(deduction_tx.mapped('amount'))
        net_pay = earning_total + tax_total + deduction_total

        if self.type == 'xlsx':
            if self.reports == 'tax_declaration':
                self.generate_tax_declaration()
                return {
                    'type': 'ir.actions.act_window',
                    'res_model': 'prx.payroll.report',
                    'view_mode': 'form',
                    'res_id': self.id,
                    'target': 'new',
                }
            # common styles
            section_font = Font(name="Microsoft JhengHei", size=9)
            label_font = Font(name="Microsoft JhengHei", size=8)
            center = Alignment(horizontal="center", vertical="center")
            left = Alignment(horizontal="left", vertical="center")
            right = Alignment(horizontal="right", vertical="center")
            black_color1 = PatternFill(fill_type="solid", fgColor="FF404040")
            black_color2 = PatternFill(fill_type="solid", fgColor="FF808080")
            black_color3 = PatternFill(fill_type="solid", fgColor="00C0C0C0")
            white_font = Font(color="FFFFFF", name="Microsoft JhengHei", size=9)
            wb = Workbook()
            ws = wb.active
            ws.title = "Payslip"

            # Row 1 header

            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15

            ws.merge_cells("A1:D1")
            ws["A1"].value = self.env.company.name
            ws["A1"].font = Font(size=12, bold=True)
            ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

            ws.merge_cells("F1:G1")
            ws["F1"].value = "ხელფასის უწყისი" if self.language == 'geo' else 'PAYSLIP'
            ws["F1"].fill = black_color1
            ws["F1"].font = white_font
            ws["F1"].alignment = center

            # Company address / VAT
            ws.merge_cells("B2:D2")
            ws.merge_cells("B3:D3")
            ws["A2"].value = "მისამართი:" if self.language == 'geo' else 'Address:'
            ws["A2"].font = label_font
            ws["B2"].value = self.env.company.street or ""
            ws["A3"].value = "ს/კ:" if self.language == 'geo' else 'Tax Number:'
            ws["A3"].font = label_font
            ws["B3"].value = self.env.company.vat or ""

            # Start Employee info section
            ws.merge_cells("A5:D5")
            ws.merge_cells("B6:D6")
            ws.merge_cells("B7:D7")
            ws.merge_cells("B8:D8")
            ws.merge_cells("B9:D9")

            ws["A5"].value = "თანამშრომლის ინფორმაცია" if self.language == 'geo' else 'EMPLOYEE INFORMATION'
            ws["A5"].fill = black_color1
            ws["A5"].font = white_font
            ws["A5"].alignment = center

            ws["A6"].value = "სახელი, გვარი:" if self.language == 'geo' else 'Full name:'
            ws["B6"].value = employee.get("name", "")
            ws["A7"].value = "მისამართი:" if self.language == 'geo' else 'Address:'
            ws["B7"].value = employee.get("private_street", "")
            ws["A8"].value = "ტელ:" if self.language == 'geo' else 'Phone:'
            ws["B8"].value = employee.get("phone", "")
            ws["A9"].value = "მეილი:" if self.language == 'geo' else 'Email:'
            ws["B9"].value = employee.get("work_email", "")

            ws["A6"].font = label_font
            ws["B6"].font = label_font
            ws["A7"].font = label_font
            ws["B7"].font = label_font
            ws["A8"].font = label_font
            ws["B8"].font = label_font
            ws["A9"].font = label_font
            ws["B9"].font = label_font
            ws["F6"].font = label_font
            ws["G6"].font = label_font

            ws["B2"].font = label_font

            ws["F5"].value = "პერიოდი" if self.language == 'geo' else 'PERIOD'
            ws["F5"].fill = black_color1
            ws["F5"].font = white_font
            ws["F5"].alignment = center
            ws["G5"].value = "დოკუმენტი" if self.language == 'geo' else 'DOCUMENT'
            ws["G5"].font = white_font
            ws["G5"].alignment = center
            ws["G5"].fill = black_color1
            ws["F6"].value = w.period_id.period or ""
            ws["G6"].value = w.sequence or ""
            ws["G6"].font = label_font
            ws["F6"].font = label_font

            # Earnings table header
            start = 11
            ws.merge_cells(f"A{start}:D{start}")
            ws[f"A{start}"].value = "ანაზღაურებები" if self.language == 'geo' else 'Earnings'
            ws[f"A{start}"].font = white_font
            ws[f"A{start}"].alignment = left
            ws[f"A{start}"].fill = black_color2
            ws[f"E{start}"].value = "როდ./სთ" if self.language == 'geo' else 'Qty/Hours'
            ws[f"E{start}"].alignment = center
            ws[f"E{start}"].font = white_font
            ws[f"E{start}"].fill = black_color2
            ws[f"F{start}"].value = "კოეფიციენტი" if self.language == 'geo' else 'Rate'
            ws[f"F{start}"].alignment = center
            ws[f"F{start}"].font = white_font
            ws[f"F{start}"].fill = black_color2
            ws[f"G{start}"].value = "თანხა" if self.language == 'geo' else 'Amount'
            ws[f"G{start}"].alignment = center
            ws[f"G{start}"].font = white_font
            ws[f"G{start}"].fill = black_color2

            # Earnings rows
            row = start + 1
            for r in earning_ids:
                ws.merge_cells(f"A{row}:D{row}")
                ws[f"A{row}"].value = r["code"]
                ws[f"A{row}"].font = label_font
                ws[f"A{row}"].alignment = left

                ws[f"E{row}"].value = r["qty"]
                ws[f"E{row}"].font = label_font
                ws[f"E{row}"].alignment = center

                ws[f"F{row}"].value = r["rate"]
                ws[f"F{row}"].font = label_font
                ws[f"F{row}"].alignment = center
                ws[f"G{row}"].value = r["amount"]
                ws[f"G{row}"].font = label_font
                ws[f"G{row}"].alignment = center
                row += 1

            # Earnings total
            ws.merge_cells(f"A{row}:E{row}")
            ws.merge_cells(f"F{row}:G{row}")
            ws[f"A{row}"].value = "ანაზღაურების ჯამი:" if self.language == 'geo' else 'Total Earning:'
            ws[f"A{row}"].font = Font(bold=True, italic=True, name="Microsoft JhengHei", size=9)
            ws[f"A{row}"].alignment = left
            ws[f"F{row}"].value = earning_total
            ws[f"A{row}"].fill = black_color3
            ws[f"F{row}"].fill = black_color3
            ws[f"F{row}"].alignment = center
            ws[f"A{row}"].alignment = right
            row += 2

            # Taxes section
            ws.merge_cells(f"A{row}:D{row}")
            ws[f"A{row}"].font = white_font
            ws[f"A{row}"].alignment = left
            ws[f"A{row}"].fill = black_color2
            ws[f"A{row}"].value = "გადასახადები" if self.language == 'geo' else 'Taxes'
            ws[f"E{row}"].fill = black_color2
            ws[f"F{row}"].font = white_font
            ws[f"F{row}"].value = "კოეფიციენტი" if self.language == 'geo' else 'Rate'
            ws[f"F{row}"].alignment = center
            ws[f"F{row}"].fill = black_color2
            ws[f"G{row}"].value = "თანხა" if self.language == 'geo' else 'Amount'
            ws[f"G{row}"].alignment = center
            ws[f"G{row}"].fill = black_color2
            ws[f"G{row}"].font = white_font
            row += 1
            for r in tax_ids:
                ws.merge_cells(f"A{row}:D{row}")
                ws[f"A{row}"].value = r["code"]
                ws[f"A{row}"].alignment = left
                ws[f"A{row}"].font = label_font
                ws[f"G{row}"].value = r["amount"]
                ws[f"G{row}"].alignment = center
                ws[f"G{row}"].font = label_font
                row += 1

            ws.merge_cells(f"F{row}:G{row}")
            ws.merge_cells(f"A{row}:E{row}")
            ws[f"A{row}"].value = "გადასახადების ჯამი:" if self.language == 'geo' else 'Total Tax:'
            ws[f"A{row}"].font = Font(bold=True, italic=True, name="Microsoft JhengHei", size=9)
            ws[f"A{row}"].alignment = right
            ws[f"F{row}"].value = tax_total
            ws[f"F{row}"].fill = black_color3
            ws[f"F{row}"].alignment = center
            ws[f"A{row}"].fill = black_color3

            row += 2

            # Deductions section
            ws.merge_cells(f"A{row}:D{row}")
            ws[f"A{row}"].font = white_font
            ws[f"A{row}"].alignment = left
            ws[f"A{row}"].fill = black_color2
            ws[f"A{row}"].value = "დაქვითვები" if self.language == 'geo' else 'Deductions'
            ws[f"E{row}"].fill = black_color2
            ws[f"F{row}"].font = white_font
            ws[f"F{row}"].value = "კოეფიციენტი" if self.language == 'geo' else 'Rate'
            ws[f"F{row}"].alignment = center
            ws[f"F{row}"].fill = black_color2
            ws[f"G{row}"].value = "თანხა" if self.language == 'geo' else 'Amount'
            ws[f"G{row}"].alignment = center
            ws[f"G{row}"].fill = black_color2
            ws[f"G{row}"].font = white_font
            row += 1
            for r in deduction_ids:
                ws.merge_cells(f"A{row}:D{row}")
                ws[f"A{row}"].value = r["code"]
                ws[f"A{row}"].alignment = left
                ws[f"A{row}"].font = label_font
                ws[f"G{row}"].value = r["amount"]
                ws[f"G{row}"].alignment = center
                ws[f"G{row}"].font = label_font
                row += 1

            ws.merge_cells(f"F{row}:G{row}")
            ws.merge_cells(f"A{row}:E{row}")
            ws[f"A{row}"].value = "დაქვითვების ჯამი:" if self.language == 'geo' else 'Total Deduction:'
            ws[f"A{row}"].font = Font(bold=True, italic=True, name="Microsoft JhengHei", size=9)
            ws[f"A{row}"].alignment = right
            ws[f"F{row}"].value = deduction_total
            ws[f"F{row}"].fill = black_color3
            ws[f"F{row}"].alignment = center
            ws[f"A{row}"].fill = black_color3
            row += 3

            # Net pay
            ws.merge_cells(f"F{row}:G{row}")
            ws.merge_cells(f"A{row}:E{row}")
            ws[f"A{row}"].value = "ხელზე ასაღები თანხა:" if self.language == 'geo' else 'Net Pay:'
            ws[f"A{row}"].font = Font(bold=True, italic=True, name="Microsoft JhengHei", size=9)
            ws[f"A{row}"].alignment = right
            ws[f"F{row}"].value = net_pay
            ws[f"F{row}"].fill = black_color3
            ws[f"F{row}"].alignment = center
            ws[f"A{row}"].fill = black_color3

            # write out to a BytesIO and encode
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            data = buf.read()

            if self.type == 'xlsx':
                self.file_download = base64.b64encode(data)
                self.file_name = f"{w.worker_id.name}_{w.period_id.period}_payslip.xlsx"
                return {
                    'type': 'ir.actions.act_window',
                    'res_model': 'prx.payroll.report',
                    'view_mode': 'form',
                    'res_id': self.id,
                    'target': 'new',
                }
            return None
        else:
            data = {
                'form': {
                    'company': self.env.company.name,
                    'address': self.env.company.street or '',
                    'taxNumber': self.env.company.vat or '',
                    'fullName': employee.get('name') or '',
                    'workerAddress': employee.get('private_street') or '',
                    'workerPhone': employee.get('phone') or '',
                    'workerEmail': employee.get('work_email') or '',
                    'period': w.period_id.period or '',
                    'document': w.sequence or '',
                    'earningTotal': str(earning_total),
                    'taxTotal': str(tax_total),
                    'deductionTotal': str(deduction_total),
                    'netPay': str(round(sum([earning_total, tax_total, deduction_total]), 2)),
                    'language': self.language,
                    'earning_ids': earning_ids,
                    'tax_ids': tax_ids,
                    'deduction_ids': deduction_ids,
                }
            }

            payload = base64.b64encode(json.dumps(data).encode()).decode()

            return {
                'type': 'ir.actions.act_url',
                'url': f'/report/payslip/view?data={payload}',
                'target': 'new',
            }
