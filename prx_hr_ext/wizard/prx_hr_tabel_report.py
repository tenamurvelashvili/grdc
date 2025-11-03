from odoo import models, fields, api
from odoo.exceptions import UserError
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import io, base64
from openpyxl.utils import range_boundaries, get_column_letter
from datetime import date
import calendar


class PRXHRTabelReport(models.TransientModel):
    _name = 'prx.hr.tabel.report'
    _description = 'Period tabel reports wizard'

    department_ids = fields.Many2many(
        'hr.department', string="დეპარტამენტი"
    )
    date = fields.Date(
        string="შედგენის თარიღი", default=fields.Date.context_today
    )

    year = fields.Integer(
        string="წელი",
        required=True,
        default=lambda self: date.today().year,
    )
    month = fields.Selection(
        [(str(i), calendar.month_name[i]) for i in range(1, 13)],
        string="თვე",
        required=True,
        default=lambda self: str(date.today().month),
    )
    period_start = fields.Date(
        string="პერიოდის დაწყება",
        compute="_onchange_month_year",
    )
    period_end = fields.Date(
        string="პერიოდის დასრულება",
        readonly=True,
        compute="_onchange_month_year",
    )

    leader = fields.Many2one('hr.employee',string='ქვედანაყოფის ხელმძღვანელი')
    responsible_person = fields.Many2one('hr.employee',string='პასუხისმგებელი პირი',default=lambda self: self.env.user.employee_id)

    @api.onchange('year', 'month')
    def _onchange_month_year(self):
        if self.year and self.month:
            m = int(self.month)
            self.period_start = date(self.year, m, 1)
            last_day = calendar.monthrange(self.year, m)[1]
            self.period_end = date(self.year, m, last_day)

    file_download = fields.Binary("File", readonly=True)
    file_name = fields.Char("Filename")

    def action_generate_tabel_report(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "ტაბელი"
        ws.sheet_view.showGridLines = False

        header_font = Font(name="Sylfaen", size=14, bold=True)
        header_font1 = Font(name="Sylfaen", size=10, bold=True)
        header_font2 = Font(name="Sylfaen", size=6, bold=True)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        right = Alignment(horizontal='right', vertical='center', wrap_text=True)
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        def merge_cell(range_str, value=None, font=None, align=None, border=None, width=None, height=None):
            ws.merge_cells(range_str)

            min_col, min_row, max_col, max_row = range_boundaries(range_str)

            if width is not None:
                for col_idx in range(min_col, max_col + 1):
                    col_letter = get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = width

            if height is not None:
                for row_idx in range(min_row, max_row + 1):
                    ws.row_dimensions[row_idx].height = height
            for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                                    min_col=min_col, max_col=max_col):
                for cell in row:
                    if cell.row == min_row and cell.column == min_col:
                        if value is not None:
                            cell.value = value
                        if font is not None:
                            cell.font = font
                        if align is not None:
                            cell.alignment = align
                    if border is not None:
                        cell.border = border

        # -- header block --
        merge_cell("A1:AN1", "დანართი N2", font=header_font, align=right, border=border)
        merge_cell("A2:AN2", "სამუშაო დროის აღრიცხვის ფორმა", font=header_font, align=center, border=border)

        company = self.env.company
        merge_cell("A3:C3", "ორგანიზაციის დასახელება", border=border, font=header_font1, align=center)
        merge_cell("D3:AN3", company.name, align=left, border=border, font=header_font1)

        merge_cell("A4:C4", "საიდენტიფიკაციო კოდი", border=border, font=header_font1, align=center)
        merge_cell("D4:AN4", company.vat or '', align=left, border=border, font=header_font1)

        merge_cell("A5:C5", "სტრუქტურული ერთეული", border=border, font=header_font1, align=center)
        depts = ", ".join(self.department_ids.mapped('name'))
        merge_cell("D5:AN5", depts or '—', align=left, border=border, font=header_font1)

        merge_cell("A6:C6", "შედგენის თარიღი", border=border, font=header_font1, align=center)
        merge_cell("D6:E6", self.date.strftime("%d.%m.%Y"), align=left, border=border, font=header_font1)
        merge_cell("F6:AN6", align=left, border=border)

        merge_cell("A7:C7", "საანგარიშო პერიოდი", border=border, font=header_font1, align=center)
        period_start = f"{self.period_start.strftime('%d.%m.%Y')} – დან"
        period_end = f"{self.period_end.strftime('%d.%m.%Y')} – მდე"
        ws["D7"] = period_start
        ws["E7"] = period_end
        ws["D7"].alignment = center
        ws["E7"].alignment = center
        ws["E7"].border = border
        ws["D7"].border = border
        ws["D7"].font = header_font1
        ws["E7"].font = header_font1

        merge_cell(
            "F6:AN7",
            ''
        )

        merge_cell(
            "A8:AN9",
            ''
        )

        row_start = 10
        row_end = 14

        merge_cell(
            f"A{row_start}:A{row_end}",
            "გვარი, სახელი",
            font=header_font2,
            align=center,
            border=border,
            height=30
        )
        merge_cell(
            f"B{row_start}:B{row_end}",
            "პირადი ნომერი/ტაბელის ნომერი",
            font=header_font2,
            align=center,
            border=border,
            height=30
        )
        merge_cell(
            f"C{row_start}:C{row_end}",
            "თანამდებობა (სპეციალობა, პროფესია)",
            font=header_font2,
            align=center,
            border=border,
            height=30
        )

        merge_cell(
            f"D10:AH10",
            "აღნიშვნები სამუშაოზე გამოცხადების/არგამოცხადების შესახებ თარიღების მიხედვით თვის განმავლობაში",
            font=header_font2,
            align=center,
            border=border,
            height=15
        )
        merge_cell(
            f"AI10:AN10",
            "სულ ნამუშევარი თვის განმავლობაში",
            font=header_font2,
            align=center,
            border=border,
        )

        merge_cell("AJ11:AN11",
                   value="საათი",
                   font=header_font2,
                   align=center,
                   border=border,
                   height=20)

        def generate_month_header():
            """
            თვეების დღეები და ქოლუმნები
            """
            header_specs = []

            first_day_col = 4
            day_start_row, day_end_row = 11, 14
            for i in range(31):
                col = first_day_col + i
                letter = get_column_letter(col)
                header_specs.append((
                    f"{letter}{day_start_row}:{letter}{day_end_row}",
                    str(i + 1)
                ))

            extra = [
                ("AI11:AI14", "დღე", {}),
                ("AJ12:AJ14", "ჯამი", {}),
                ("AK13:AK14", "ზე განაკვეთი", {}),
                ("AL13:AL14", "ღამე", {}),
                ("AM13:AM14", "დასვენება/ უქმე დღეებში ნამუშევარი საათების ჯამური რაოდენობა (თვე)", {}),
                ("AN13:AN14", "სხვა (საჭიროების შემთხვევაში)", {}),
                ("AK12:AN12", "მათ შორის", {"height": 10}),
            ]
            for rng, label, opts in extra:
                header_specs.append((rng, label, opts))

            for spec in header_specs:
                rng, label = spec[0], spec[1]
                opts = spec[2] if len(spec) > 2 else {}
                merge_cell(
                    rng,
                    label,
                    font=header_font2,
                    align=center,
                    border=border,
                    **opts
                )

        generate_month_header()

        def create_15_line():
            left_values = {1: 1, 2: 2, 3: 3}
            for col_idx, val in left_values.items():
                letter = get_column_letter(col_idx)
                cell = ws[f"{letter}15"]
                cell.value = val
                cell.alignment = center
                cell.border = border
                cell.font = header_font2

            for offset, val in enumerate(range(5, 11)):
                col_idx = 35 + offset
                letter = get_column_letter(col_idx)
                cell = ws[f"{letter}15"]
                cell.value = val
                cell.alignment = center
                cell.border = border
                cell.font = header_font2

            merge_cell(
                "D15:AH15",
                value=4,
                font=header_font2,
                align=center,
                border=border,
                height=20
            )

        create_15_line()

        month_geo = {
            '1': 'იანვარი',
            '2': 'თებერვალი',
            '3': 'მარტი',
            '4': 'აპრილი',
            '5': 'მაისი',
            '6': 'ივნისი',
            '7': 'ივლისი',
            '8': 'აგვისტო',
            '9': 'სექტემბერი',
            '10': 'ოქტომბერი',
            '11': 'ნოემბერი',
            '12': 'დეკემბერი',
        }

        def get_employee_work_data(employee):
            year = self.year
            month = month_geo[str(self.month)]

            calendar = self.env['prx.organisation.calendar'].search([
                ('company_id', '=', self.env.company.id),
                ('year', '=', year),
                ('schedule_type_id', '=', employee.resource_calendar_id.id),
            ])
            if not calendar:
                raise UserError(f'კალენდარი არ მოიძებნა {employee.name} - {employee.resource_calendar_id.name}')
            details = calendar.calendar_details_id.filtered(
                lambda d: d.month == month
            )
            work_days = details.filtered(lambda d: d.status == 'open' and not d.holiday)
            holidays = details.filtered(lambda d: d.holiday)

            def get_worked_hours(date):
                attendance = self.env['resource.calendar.attendance'].get_workday_hours(
                    employee.resource_calendar_id,
                    date
                )
                return attendance

            def calc_status_doc(status):
                leave_type_dict = {
                    'Open': f'{calendar.schedule_type_id.hours_per_day}:00',
                    'NoWorkDay': 'დ',
                    'PaidLive': 'ა / შ',
                    'UnpaidLive': 'უ / შ',
                    'PaidMaternityLeave': 'დ / შ - ა',
                    'UnPaidMaternityLeave': 'დ / შ - უ',
                    'SickLive': 'ს / ფ',
                    'Absence': 'გ',
                }

                return leave_type_dict[status]

            employees_data = []
            sum_work_days = 0.0
            sum_work_hours = 0.0
            for data in details:
                day_of_month = data.date.day
                obj = {'Date': data.date, 'DayMonth': day_of_month, 'Status': None}
                if data.holiday:
                    obj['Status'] = calc_status_doc('NoWorkDay')
                    employees_data.append(obj)
                    continue
                status = data.status
                if status == 'open' and not data.holiday:
                    hr_leve = self.env['hr.leave'].search(
                        [('request_date_from', '<=', data.date),
                         ('request_date_to', '>=', data.date),
                         ('employee_id', '=', employee.id),
                         ('state', '=', 'validate')], limit=1)
                    if hr_leve:
                        obj['Status'] = hr_leve.holiday_status_id.prx_time_off_code_id.code
                        employees_data.append(obj)
                    else:

                        def hhmm_to_float(hhmm: str) -> float:
                            hours_str, mins_str = hhmm.split(':')
                            hours1 = int(hours_str)
                            minutes = int(mins_str)
                            return hours1 + minutes / 60.0

                        hours_workday = get_worked_hours(data.date)
                        obj['Status'] = hours_workday
                        employees_data.append(obj)
                        to_float_hours = hhmm_to_float(hours_workday)
                        if to_float_hours:
                            sum_work_days += 1
                            sum_work_hours += to_float_hours

                elif status == 'closed':
                    obj['Status'] = calc_status_doc('NoWorkDay')
                    employees_data.append(obj)

            return employees_data, sum_work_days, sum_work_hours

        def generate_employee_table():
            employees = self.env['hr.employee'].search([])
            if self.department_ids:
                employees = employees.search([('department_id', 'in', self.department_ids.ids)])

            current_row = 16
            for emp in employees:
                cell_a = ws.cell(row=current_row, column=1, value=emp.name)
                cell_a.font = header_font2
                cell_a.alignment = left
                cell_a.border = border

                cell_b = ws.cell(row=current_row, column=2, value=emp.identification_id or emp.barcode or '')
                cell_b.font = header_font2
                cell_b.alignment = center
                cell_b.border = border

                cell_c = ws.cell(row=current_row, column=3, value=emp.job_id.name or '')
                cell_c.font = header_font2
                cell_c.alignment = left
                cell_c.border = border

                employee_work_data, sum_work_days, sum_work_hours = get_employee_work_data(emp)
                rows_len = 0
                if len(employee_work_data) < 31:
                    rows_len = 31 - len(employee_work_data)

                for offset, day_info in enumerate(employee_work_data, start=0):
                    col_idx = 4 + offset
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.value = day_info['Status']
                    cell.border = border
                    cell.alignment = center

                # ცარიელი თუ არის მაშინ ბლანკს გავხდი
                for offset in range(rows_len):
                    col_idx = 4 + len(employee_work_data) + offset
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.border = border
                    cell.alignment = center

                # ნამუშევარი დღეების ჯამი
                last_index = 4 + 31
                cell = ws.cell(row=current_row, column=last_index)
                cell.border = border
                cell.alignment = center
                cell.value = sum_work_days

                # ნამუშევარი საატების ჯამი
                if sum_work_days:
                    cell = ws.cell(row=current_row, column=last_index+1)
                    cell.border = border
                    cell.alignment = center
                    cell.value = sum_work_hours

                # ბოლო 4 სრიქონის შევსება მხოლოდ ბორდერებით ჯერჯერობით
                for rec in range(4):
                    cell = ws.cell(row=current_row, column=(last_index+2) + rec)
                    cell.border = border
                    cell.alignment = center


                ws.row_dimensions[current_row].height = 18

                current_row += 1
            return current_row

        row_after_add_data = generate_employee_table()

        def generate_organisation_structure():

            last_row = row_after_add_data
            end_last_row = last_row + 1

            merge_cell(
                f"A{last_row}:AN{end_last_row}",
                "",
                font=header_font1,
                align=center,
            )
            merge_cell(
                f"A{last_row + 2}:C{last_row + 3}",
                "ორგანიზაციის/სტრუქტურული ქვედანაყოფის ხელმძღვანელი",
                font=header_font1,
                align=center,
                border=border,
                height=50
            )

            merge_cell(
                f"A{last_row + 5}:C{last_row + 6}",
                "სამუშაო დროის აღრიცხვის ფორმის შედგენაზე პასუხისმგებელი პირი",
                font=header_font1,
                align=center,
                border=border,
                height=50
            )

            merge_cell(
                f"D{last_row + 2}:G{last_row + 2}",
                self.leader.name,
                font=header_font1,
                align=center,
                border=border,
                width=10
            )
            merge_cell(
                f"D{last_row + 5}:G{last_row + 5}",
                self.responsible_person.name,
                font=header_font1,
                align=center,
                border=border,
                width=10
            )
            merge_cell(
                f"D{last_row + 3}:G{last_row + 3}",
                "გვარი, სახელი",
                font=header_font1,
                align=center,
                border=border,
                # width=10,
                height=20
            )

            merge_cell(
                f"D{last_row + 6}:G{last_row + 6}",
                "გვარი, სახელი",
                font=header_font1,
                align=center,
                border=border,
                # width=10,
                height=20
            )

            merge_cell(
                f"I{last_row + 2}:J{last_row + 2}",
                "",
                font=header_font1,
                align=center,
                border=border,
                width=10
            )
            merge_cell(
                f"I{last_row + 5}:J{last_row + 5}",
                "",
                font=header_font1,
                align=center,
                border=border,
                width=10
            )
            merge_cell(
                f"I{last_row + 3}:J{last_row + 3}",
                "ხელმოწერა",
                font=header_font1,
                align=center,
                border=border,
                width=10
            )
            merge_cell(
                f"I{last_row + 6}:J{last_row + 6}",
                "ხელმოწერა",
                font=header_font1,
                align=center,
                border=border,
                width=10
            )
            return last_row + 6

        last_rw = generate_organisation_structure()

        def create_last_descriptions():
            header_font3 = Font(name="Sylfaen", size=11, bold=False)
            header_data = Font(name="Sylfaen", size=9, bold=True)
            header_data2 = Font(name="Sylfaen", size=9, bold=False)

            start_row = last_rw + 3

            merge_cell(
                f"A{start_row}:C{start_row}",
                "პირობითი აღნიშვნები",
                font=header_font3,
                align=center,
                border=border,
                width=10
            )

            labels = [
                "დ",
                "ა / შ",
                "უ / შ",
                "დ / შ - ა",
                "დ / შ - უ",
                "ს / ფ",
                "გ",
            ]
            start_row = start_row + 1
            for ind, text in enumerate(labels):
                row = f"A{start_row + ind}"
                row2 = f"B{start_row + ind}"
                ws[row] = text
                ws[row].alignment = center
                ws[row].border = border
                ws[row].font = header_data
                # row2
                ws[row2] = ''
                ws[row2].alignment = center
                ws[row2].border = border
                ws[row2].font = header_data
            # row3
            more_labels = [
                "დასვენება / უქმე დღეები",
                "ანაზღაურებადი შვებულება",
                "ანაზღაურების გარეშე შვებულება",
                "დეკრეტული შვებულება - ანაზღაურებადი",
                "დეკრეტული შვებულება - ანაზღაურების გარეშე",
                "საავადმყოფო ფურცელი",
                "გაცდენა",
            ]

            for indx, text in enumerate(more_labels):
                row = f"C{start_row + indx}"
                ws[row] = text
                ws[row].alignment = center
                ws[row].border = border
                ws[row].font = header_data2

        create_last_descriptions()

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        self.file_download = base64.b64encode(bio.read())
        self.file_name = "Period_Tabel.xlsx"

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }
